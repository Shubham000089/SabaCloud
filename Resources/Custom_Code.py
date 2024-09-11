# Imports for Excel file program
from openpyxl import Workbook, load_workbook
import os
from contextlib import closing
from openpyxl.utils import get_column_letter
import openpyxl
# Imports for calendar program
from datetime import datetime, timedelta
import sys


# Imports for custom mail
import random
import string


try:
    from robot.libraries.BuiltIn import BuiltIn
    from robot.libraries.BuiltIn import _Misc
    import robot.api.logger as logger
    from robot.api.deco import keyword

    ROBOT = False
except Exception:
    ROBOT = False


file_name = "SabaCloud.xlsx"


def get_current_date_time():
    now = datetime.now()
    formatted_date_time = now.strftime("%d-%m-%Y %H:%M:%S")
    return formatted_date_time


def give_value():
    with closing(load_workbook(filename=file_name)) as wb:
        ws = wb.active
        # from here change
        col_letter = get_column_letter(1)
        max_col_row = len([cell for cell in ws[col_letter] if cell.value])
        return max_col_row


def read_test_data(cell):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_name)

    # Select the specified sheet
    sheet = wb['Sheet1']

    # Get the value of the specified cell
    cell_value = sheet[cell].value

    # Close the workbook
    wb.close()

    return cell_value


# def add_value(cell_cords_1, value):
#     with closing(load_workbook(filename=file_name)) as wb:
#         ws = wb.active
#         # from here change
#         col_letter = get_column_letter(9)
#         print(col_letter)
#         max_col_row = len([cell for cell in ws[col_letter] if cell.value])
#         max_col_row_1 = max_col_row + 1
#         # up to here
#         cell_cords = f"{cell_cords_1}{max_col_row_1}"
#         ws[cell_cords] = value
#         wb.save(file_name)

def add_value(cell_cords_1, value, file_name):
    with closing(load_workbook(filename=file_name)) as wb:
        ws = wb.active
        # from here change
        col_letter = get_column_letter(10)
        print(col_letter)
        max_col_row = len([cell for cell in ws[col_letter] if cell.value])
        max_col_row_1 = max_col_row + 1
        # up to here
        cell_cords = f"{cell_cords_1}{max_col_row_1}"
        ws[cell_cords] = value
        wb.save(file_name)


def gen_random_string(length=3):
    characters = string.ascii_letters + string.digits  # A-Z, a-z, 0-9
    random_string = ''.join(random.choice(characters) for _ in range(length))
    return random_string


def save_file(file_path):
    pyautogui.hotkey('ctrl', 's')
    pyautogui.sleep(2)
    pyautogui.typewrite(file_path)
    for _ in range(3):
        pyautogui.press('tab')
        pyautogui.sleep(0.5)
    pyautogui.press('enter')
    pyautogui.sleep(3)

def start_virtual_display():
    display = Display(visible=True, size=(1920, 1080))
    display.start()
    return display

def stop_virtual_display(display):
    display.stop()

if __name__ == "__main__":
    file_path = sys.argv[1]
    save_file(file_path)