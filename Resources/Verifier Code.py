import os
import openpyxl
import Levenshtein

# Specify the folder path
folder_path = 'C:\SabaCloud_Reports\WHITE'
file_path = '../Extracted Excel/WHITE.xlsx'


def list_filenames_with_extensions(folder_path):
    # Get a list of all files and directories in the specified folder
    files = os.listdir(folder_path)

    # Filter out only files (not directories) and return them with their extensions
    filenames_with_extensions = [f for f in files if os.path.isfile(os.path.join(folder_path, f))]

    return filenames_with_extensions


def get_downloaded_names(file_path):
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)

    # Select the active sheet (or you can use workbook['SheetName'] to specify the sheet)
    sheet = workbook.active

    downloaded_names = []

    # Loop through the rows in the sheet, starting from row 2 to skip the header
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=9, max_col=10, values_only=True):
        name = row[0]  # Column I (index 0 for I)
        status = row[1]  # Column J (index 1 for J)

        # Check if the status is "Downloaded"
        if status == "Downloaded":
            downloaded_names.append(name)

    # Return the final list of names
    return downloaded_names


def get_not_downloaded_names(file_path):
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)

    # Select the active sheet (or you can use workbook['SheetName'] to specify the sheet)
    sheet = workbook.active

    downloaded_names = []

    # Loop through the rows in the sheet, starting from row 2 to skip the header
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=9, max_col=10, values_only=True):
        name = row[0]  # Column I (index 0 for I)
        status = row[1]  # Column J (index 1 for J)

        # Check if the status is "Downloaded"
        if status == "Not Downloaded":
            downloaded_names.append(name)

    # Return the final list of names
    return downloaded_names


def mix_downloaded_names(file_path):
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)

    # Select the active sheet (or you can use workbook['SheetName'] to specify the sheet)
    sheet = workbook.active

    downloaded_names = []

    # Loop through the rows in the sheet, starting from row 2 to skip the header
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=9, max_col=10, values_only=True):
        name = row[0]  # Column I (index 0 for I)
        status = row[1]  # Column J (index 1 for J)

        # Check if the status is "Downloaded"
        if status == "Not Downloaded" or status == "Downloaded":
            downloaded_names.append(name)

    # Return the final list of names
    return downloaded_names


def update_excel_with_presence(file_path, folder_path):
    """Check status from Excel and update presence in column K based on filenames from folder."""
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active  # Select the active sheet

    # Get list of filenames from the folder
    folder_filenames = list_filenames_with_extensions(folder_path)

    # Loop through the rows in the sheet starting from row 2 to skip the header
    for row in range(2, sheet.max_row + 1):
        status = sheet.cell(row=row, column=10).value  # Column J (Status)
        name = sheet.cell(row=row, column=9).value     # Column I (Name)

        if status in ["Downloaded", "Not Downloaded"]:
            # Check if the name exists in the list of filenames from the folder
            if name in folder_filenames:
                sheet.cell(row=row, column=11).value = "Present"  # Column K
            else:
                sheet.cell(row=row, column=11).value = "Not Present"  # Column K

    # Save the workbook with changes
    workbook.save(file_path)


def get_new_names_from_folder(excel_file_path, folder_path):
    """Compare the 'Downloaded' names from Excel with filenames in the folder, return list of new names."""
    # Get the list of 'Downloaded' names from Excel
    downloaded_names = get_downloaded_names(excel_file_path)

    # Get the list of filenames from the folder
    folder_filenames = list_filenames_with_extensions(folder_path)

    # Find names that are in the folder but not in the downloaded list
    new_names = [name for name in folder_filenames if name not in downloaded_names]

    return new_names


def calculate_matching_rate(name1, name2):
    """Calculate the matching rate between two strings using Levenshtein distance."""
    distance = Levenshtein.distance(name1, name2)
    max_len = max(len(name1), len(name2))
    matching_rate = (1 - distance / max_len) * 100  # Convert to percentage
    return round(matching_rate, 2)


def update_excel_with_matching_rate(file_path, folder_path):
    """Update Excel file with matching rate for 'Not Downloaded' and 'Not Present' names in column L."""
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active  # Select the active sheet

    # Get list of filenames from the folder
    folder_filenames = list_filenames_with_extensions(folder_path)

    # Loop through rows and perform matching
    for row in range(2, sheet.max_row + 1):
        status_j = sheet.cell(row=row, column=10).value  # Column J (Status)
        status_k = sheet.cell(row=row, column=11).value  # Column K (Presence)
        name = sheet.cell(row=row, column=9).value  # Column I (Name)

        if status_j == "Not Downloaded" and status_k == "Not Present":
            # Compare the name from Excel with filenames from folder
            best_match_rate = 0
            for folder_name in folder_filenames:
                match_rate = calculate_matching_rate(name, folder_name)
                if match_rate > best_match_rate:
                    best_match_rate = match_rate

            # Write the best matching rate in column L
            sheet.cell(row=row, column=12).value = best_match_rate  # Column L

    # Save the workbook with changes
    workbook.save(file_path)


# Call the function and print the filenames with extensions
filenames = list_filenames_with_extensions(folder_path)
print("Count of attachment from folders:", len(filenames))
# print(filenames)

downloaded_names_list = get_downloaded_names(file_path)
print("The count of downloaded attachment is", len(downloaded_names_list))
# # print(downloaded_names_list)
#
not_downloaded_names_list = get_not_downloaded_names(file_path)
print("The count of not downloaded attachment is", len(not_downloaded_names_list))
# # print(not_downloaded_names_list)
#
mix_downloaded_names_list = mix_downloaded_names(file_path)
print("The count of downloaded and non downloaded attachment is", len(mix_downloaded_names_list))
# print(not_downloaded_names_list)

update_excel_with_presence(file_path, folder_path)
print("Excel file updated successfully.")

# new_names_list = get_new_names_from_folder(file_path, folder_path)
# print("Count of list is: ", len(new_names_list))
# print("New names found in folder but not in Excel 'Downloaded' list:", new_names_list)

# update_excel_with_matching_rate(file_path, folder_path)
# print("Excel file updated with matching rates in column L.")